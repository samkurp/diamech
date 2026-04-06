"""
Генерация протоколов и ZIP архивов
"""
import os
import io
import zipfile
import tempfile
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from .config import Config
from .image_utils import compress_image, allowed_file
from .database import SupabaseDB


def generate_folder_name(data):
    """Генерирует имя папки для сохранения"""
    machine_type = data.get('machineType', '').strip()
    lifting_capacity = data.get('liftingCapacity', '').strip()
    serial_number = data.get('serialNumber', 'unknown').strip().replace(' ', '_')
    return f"ML_{machine_type}{lifting_capacity}№{serial_number}"


def transliterate_machine_type(machine_type):
    """Транслитерирует тип станка для имени ZIP файла"""
    translit_map = {
        'В': 'B',
        'ВТ': 'BT',
        'ВМ': 'BM',
        'СП': 'SP',
        'ДБС': 'DBS'
    }
    return translit_map.get(machine_type, machine_type)


def generate_protocol_filename(data):
    """Генерирует имя файла протокола (кириллица)"""
    machine_type = data.get('machineType', '').strip()
    lifting_capacity = data.get('liftingCapacity', '').strip()
    serial_number = data.get('serialNumber', 'unknown').strip().replace(' ', '_')
    return f"{machine_type}{lifting_capacity}№{serial_number}.xlsx"


def generate_zip_filename(data):
    """Генерирует имя ZIP файла (латиница)"""
    machine_type = data.get('machineType', '').strip()
    lifting_capacity = data.get('liftingCapacity', '').strip()
    serial_number = data.get('serialNumber', 'unknown').strip().replace(' ', '_')

    # Транслитерируем тип станка
    machine_type_lat = transliterate_machine_type(machine_type)

    return f"{machine_type_lat}{lifting_capacity}№{serial_number}.zip"


def generate_info_content(data):
    """Генерирует содержимое информационного файла README.txt"""
    return f"""
ПРОТОКОЛ

=====================================


1. ОСНОВНАЯ ИНФОРМАЦИЯ
-----------------------
Тип станка:            {data.get('machineType', 'Не указан')}
Грузоподъемность:      {data.get('liftingCapacity', 'Не указана')}
Заводской номер:       {data.get('serialNumber', 'Не указан')}
Вид работ:            {data.get('workType', 'Не указан')}
Заказчик:             {data.get('customer', 'Не указан')}
Статус:               {data.get('machineStatus', 'Сборка')}
Плановая отгрузка:    {data.get('shippingDate', 'Не указана')}

2. ПРИВОДНАЯ СИСТЕМА
--------------------
Тип привода:          {data.get('driveType', 'Не указан')}
Номер привода:        {data.get('driveNumber', 'Не указан')}
Тормозной резистор:   {data.get('brakeResistor', 'Не указан')}
Кол-во резисторов:    {data.get('resistorCount', '-')}

3. ЭЛЕКТРОДВИГАТЕЛЬ
-------------------
Тип двигателя:        {data.get('electricMotor', 'Не указан')}
Мощность:             {data.get('EnginePower', 'Не указана')}
Номер двигателя:      {data.get('motorNumber', 'Не указан')}

4. ДАТЧИКИ
-----------
Датчик угла:          {data.get('angleSensor', 'Не указан')}
Номер датчика угла:   {data.get('angleSensorNumber', 'Не указан')}
Номер отметчика:      {data.get('speedSensorNumber', 'Не указан')}

5. ДАТЧИКИ ВИБРАЦИИ
-------------------
Левый датчик:         {data.get('leftVibrationSensor', 'Не указан')}
Чувствительность:     {data.get('leftSensitivity', '-')}
Номер датчика:        {data.get('leftSensorNumber', 'Не указан')}

Правый датчик:        {data.get('rightVibrationSensor', 'Не указан')}
Чувствительность:     {data.get('rightSensitivity', '-')}
Номер датчика:        {data.get('rightSensorNumber', 'Не указан')}

6. ИЗМЕРИТЕЛЬНОЕ ОБОРУДОВАНИЕ
-----------------------------
Измерительный прибор: {data.get('measuringDevice', 'Не указан')}
Номер прибора:        {data.get('measuringDeviceNumber', 'Не указан')}
Блок обработки:       {data.get('signalProcessor', 'Не указан')}
Номер блока:          {data.get('signalProcessorNumber', 'Не указан')}


8. ПРИМЕЧАНИЯ
-------------
{data.get('notes', 'Нет')}

=====================================
Конец протокола
    """.strip()


def fill_excel_template(template_path, data):
    """
    Заполняет шаблон Excel данными
    
    Args:
        template_path: путь к шаблону Excel
        data: данные для заполнения
    
    Returns:
        Workbook с заполненными данными
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # Маппинг полей на ячейки Excel
    mapping = {
        'workType': 'I1',
        'machineType': 'C3',
        'liftingCapacity': 'D3',
        'serialNumber': 'J3',
        'driveType': 'C5',
        'driveNumber': 'F5',
        'brakeResistor': 'E7',
        'resistorCount': 'H7',
        'electricMotor': 'D9',
        'motorNumber': 'G9',
        'EnginePower': 'K9',
        'angleSensor': 'D11',
        'angleSensorNumber': 'G11',
        'speedSensorNumber': 'K11',
        'leftVibrationSensor': 'D15',
        'leftSensitivity': 'G15',
        'leftSensorNumber': 'I15',
        'rightVibrationSensor': 'D16',
        'rightSensitivity': 'G16',
        'rightSensorNumber': 'I16',
        'measuringDevice': 'E18',
        'measuringDeviceNumber': 'G18',
        'signalProcessor': 'E20',
        'signalProcessorNumber': 'G20',
        'notes': 'A37'
    }

    for field, cell in mapping.items():
        if field in data and data[field]:
            ws[cell] = data[field]

    return wb


def generate_protocol(data, request_files=None, draft_id=None):
    """
    Генерирует полный пакет: протокол Excel + все изображения
    и возвращает ZIP архив в виде байтов
    
    Args:
        data: данные формы
        request_files: файлы изображений из запроса
        draft_id: ID черновика для загрузки дополнительных данных из БД
    
    Returns:
        tuple (zip_data, zip_filename) или dict с ошибкой
    """
    try:
        if not os.path.exists(Config.TEMPLATE_PATH):
            return {'success': False, 'error': f'Шаблон {Config.TEMPLATE_PATH} не найден'}

        # Создаем временную директорию
        with tempfile.TemporaryDirectory() as temp_dir:
            folder_name = generate_folder_name(data)
            protocol_filename = generate_protocol_filename(data)

            # 1. Генерируем протокол Excel из шаблона
            wb = fill_excel_template(Config.TEMPLATE_PATH, data)
            
            # Сохраняем протокол
            protocol_path = os.path.join(temp_dir, protocol_filename)
            wb.save(protocol_path)
            wb.close()

            # 2. Создаем папку для изображений
            images_dir = os.path.join(temp_dir, f"{folder_name}_images")
            os.makedirs(images_dir, exist_ok=True)

            # 3. Сохраняем изображения из формы
            saved_images = []
            if request_files:
                for img in request_files:
                    if img and allowed_file(img.filename):
                        filename = secure_filename(img.filename)
                        img.seek(0)
                        img_data = img.read()

                        # Сжимаем изображение
                        compressed_data = compress_image(
                            img_data,
                            max_size=Config.IMAGE_MAX_SIZE,
                            quality=Config.IMAGE_QUALITY
                        )

                        # Сохраняем во временную папку
                        img_path = os.path.join(images_dir, filename)
                        with open(img_path, 'wb') as f:
                            f.write(compressed_data)
                        saved_images.append(filename)
                        print(f"📸 Сохранено изображение из формы: {filename}")

            # 4. Если есть draft_id, загружаем дополнительные изображения из БД
            if draft_id:
                draft = SupabaseDB.get_draft(draft_id)
                if draft and draft.get('image_files'):
                    for img_filename in draft['image_files']:
                        if img_filename not in saved_images:
                            image_data = SupabaseDB.get_image(draft_id, img_filename)
                            if image_data:
                                img_bytes, _ = image_data
                                img_path = os.path.join(images_dir, img_filename)
                                with open(img_path, 'wb') as f:
                                    f.write(img_bytes)
                                saved_images.append(img_filename)
                                print(f"📸 Сохранено изображение из БД: {img_filename}")

            # 5. Создаем информационный файл README.txt
            info_content = generate_info_content(data)
            info_path = os.path.join(temp_dir, "README.txt")
            with open(info_path, 'w', encoding='utf-8') as f:
                f.write(info_content)

            # 6. Создаем ZIP архив
            zip_filename = generate_zip_filename(data)
            zip_path = os.path.join(temp_dir, zip_filename)

            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                # Добавляем протокол Excel
                zf.write(protocol_path, protocol_filename)

                # Добавляем README с информацией
                zf.write(info_path, "README.txt")

                # Добавляем все изображения в папку images
                for img_file in os.listdir(images_dir):
                    img_full_path = os.path.join(images_dir, img_file)
                    zf.write(img_full_path, f"images/{img_file}")

            print(f"✅ ZIP архив создан: {zip_filename}, размер: {os.path.getsize(zip_path)} байт")

            # 7. Читаем ZIP файл в память
            with open(zip_path, 'rb') as f:
                zip_data = f.read()

            return {'success': True, 'data': zip_data, 'filename': zip_filename}

    except Exception as e:
        print(f"❌ Ошибка генерации пакета: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'Ошибка генерации протокола: {str(e)}'}
