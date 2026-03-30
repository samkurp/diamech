"""
Flask приложение для управления производственными данными станков
Production-ready версия с модульной архитектурой
"""
import os
from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

# Импортируем конфигурацию
from app.config import config_by_name, Config

# Импортируем утилиты
from app.utils import (
    setup_logger,
    compress_image,
    extract_text_from_file,
    allowed_file,
    generate_folder_name,
    generate_protocol_filename,
    generate_zip_filename
)

# Импортируем сервисы
from app.services.supabase_service import SupabaseService


def create_app(config_name=None):
    """Фабрика приложений Flask"""
    
    if config_name is None:
        config_name = os.environ.get('FLASK_ENV', 'development')
    
    app = Flask(__name__, static_folder='static')
    
    # Загружаем конфигурацию
    config_class = config_by_name.get(config_name, config_by_name['default'])
    app.config.from_object(config_class)
    
    # Инициализация CORS
    CORS(app)
    
    # Настройка логгера
    logger = setup_logger(app)
    
    # Инициализация сервиса Supabase
    supabase_service = SupabaseService()
    
    # Регистрируем маршруты
    register_routes(app, supabase_service, logger)
    
    # Регистрируем обработчики ошибок
    register_error_handlers(app)
    
    logger.info(f'✅ Приложение инициализировано в режиме: {config_name}')
    
    return app


def register_routes(app, supabase_service, logger):
    """Регистрация всех маршрутов приложения"""
    
    # ========== HEALTH CHECK ==========
    @app.route('/api/health', methods=['GET'])
    def health_check():
        """Проверка состояния сервера"""
        return jsonify({
            'success': True,
            'status': 'ok',
            'service': 'machine-protocol-generator',
            'supabase_configured': supabase_service.is_connected(),
            'template_exists': os.path.exists(Config.TEMPLATE_PATH),
            'image_compression': {
                'enabled': True,
                'max_size': Config.IMAGE_MAX_SIZE,
                'quality': Config.IMAGE_QUALITY,
                'format': Config.IMAGE_FORMAT
            },
            'timestamp': __import__('datetime').datetime.now().isoformat()
        })
    
    # ========== DRAFTS ==========
    @app.route('/api/drafts', methods=['GET'])
    def get_drafts():
        """Получает список черновиков"""
        try:
            status_filter = request.args.get('status')
            drafts = supabase_service.get_all_drafts(status_filter)
            
            return jsonify({
                'success': True,
                'drafts': drafts,
                'total': len(drafts),
                'filter': status_filter
            })
        except Exception as e:
            logger.error(f"Ошибка в /api/drafts: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>', methods=['GET'])
    def get_draft(draft_id):
        """Получает конкретный черновик"""
        try:
            draft = supabase_service.get_draft(draft_id)
            
            if draft:
                return jsonify({'success': True, 'draft': draft})
            else:
                return jsonify({'success': False, 'error': 'Черновик не найден'}), 404
        except Exception as e:
            logger.error(f"Ошибка получения черновика {draft_id}: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/save-draft', methods=['POST'])
    def save_draft():
        """Сохраняет черновик со сжатыми изображениями"""
        try:
            data = request.form
            images = request.files.getlist('images')
            
            def compress_func(img_data):
                return compress_image(img_data, Config.IMAGE_MAX_SIZE, Config.IMAGE_QUALITY)
            
            success, result, error = supabase_service.save_draft(data, images, compress_func)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': 'Черновик успешно сохранен',
                    'draft_id': result['id'],
                    'draft_data': result,
                    'image_compression': {
                        'max_size': Config.IMAGE_MAX_SIZE,
                        'quality': Config.IMAGE_QUALITY
                    }
                })
            else:
                return jsonify({'success': False, 'error': result}), 400
        except Exception as e:
            logger.error(f"Ошибка сохранения черновика: {e}")
            return jsonify({'success': False, 'error': f'Ошибка сохранения: {str(e)}'}), 500
    
    @app.route('/api/drafts/<draft_id>', methods=['PUT'])
    def update_draft(draft_id):
        """Обновляет черновик"""
        try:
            data = request.form
            images = request.files.getlist('images')
            
            def compress_func(img_data):
                return compress_image(img_data, Config.IMAGE_MAX_SIZE, Config.IMAGE_QUALITY)
            
            success, result = supabase_service.update_draft(draft_id, data, images, compress_func)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': 'Черновик обновлен',
                    'draft_data': result
                })
            else:
                return jsonify({'success': False, 'error': result}), 400
        except Exception as e:
            logger.error(f"Ошибка обновления черновика: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>/customer', methods=['GET', 'POST', 'PUT'])
    def manage_customer_data(draft_id):
        """Управление данными заказчика"""
        try:
            if request.method == 'GET':
                customer_data = supabase_service.get_customer_data(draft_id)
                
                if customer_data:
                    return jsonify({'success': True, 'customer_data': customer_data})
                else:
                    return jsonify({'success': False, 'error': 'Данные заказчика не найдены'}), 404
            
            elif request.method in ['POST', 'PUT']:
                data = request.get_json()
                
                if not data:
                    return jsonify({'success': False, 'error': 'Нет данных для сохранения'}), 400
                
                success, result = supabase_service.update_customer_data(draft_id, data)
                
                if success:
                    return jsonify({
                        'success': True,
                        'message': 'Данные заказчика сохранены',
                        'draft_data': result
                    })
                else:
                    return jsonify({'success': False, 'error': result}), 400
        except Exception as e:
            logger.error(f"Ошибка управления данными заказчика: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>/images/<filename>')
    def get_draft_image(draft_id, filename):
        """Возвращает изображение"""
        try:
            image_data = supabase_service.get_image(draft_id, filename)
            
            if image_data:
                img_bytes, content_type = image_data
                return send_file(
                    __import__('io').BytesIO(img_bytes),
                    mimetype=content_type,
                    as_attachment=False,
                    download_name=filename
                )
            else:
                return jsonify({'success': False, 'error': 'Изображение не найдено'}), 404
        except Exception as e:
            logger.error(f"Ошибка получения изображения: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>/delete', methods=['POST', 'DELETE'])
    def delete_draft(draft_id):
        """Удаляет черновик"""
        try:
            confirm = request.args.get('confirm')
            
            if confirm != 'yes':
                return jsonify({
                    'success': False,
                    'error': 'Для удаления требуется подтверждение'
                }), 400
            
            success, message = supabase_service.delete_draft(draft_id)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': message,
                    'draft_id': draft_id
                })
            else:
                return jsonify({'success': False, 'error': message}), 404 if message == "Черновик не найден" else 500
        except Exception as e:
            logger.error(f"Ошибка удаления черновика: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # ========== REQUEST FILES ==========
    @app.route('/api/drafts/<draft_id>/request-text', methods=['GET'])
    def get_request_text(draft_id):
        """Получает извлеченный текст заявки"""
        try:
            if not supabase_service.client:
                return jsonify({'success': False, 'error': 'Supabase не инициализирован'}), 500
            
            response = supabase_service.client.table('request_files') \
                .select('extracted_text, has_text, filename') \
                .eq('draft_id', draft_id) \
                .execute()
            
            if response.data and response.data[0].get('has_text'):
                return jsonify({
                    'success': True,
                    'text': response.data[0]['extracted_text'],
                    'filename': response.data[0]['filename']
                })
            else:
                return jsonify({'success': False, 'error': 'Текст не найден или не извлечен'}), 404
        except Exception as e:
            logger.error(f"Ошибка получения текста заявки: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>/upload-request', methods=['POST'])
    def upload_request_file(draft_id):
        """Загружает файл заявки для черновика"""
        try:
            if 'requestFile' not in request.files:
                return jsonify({'success': False, 'error': 'Файл не передан'}), 400
            
            file = request.files['requestFile']
            if file.filename == '':
                return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
            
            success, result = supabase_service.save_request_file(draft_id, file, extract_text_from_file)
            
            if success:
                return jsonify({
                    'success': True,
                    'message': 'Файл заявки сохранен',
                    'filename': result
                })
            else:
                return jsonify({'success': False, 'error': result}), 400
        except Exception as e:
            logger.error(f"Ошибка загрузки файла заявки: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/drafts/<draft_id>/request-file', methods=['GET', 'POST', 'DELETE'])
    def manage_request_file(draft_id):
        """Управление файлом заявки"""
        try:
            if request.method == 'GET':
                file_data = supabase_service.get_request_file(draft_id)
                
                if file_data:
                    file_bytes, filename, content_type = file_data
                    return send_file(
                        __import__('io').BytesIO(file_bytes),
                        mimetype=content_type,
                        as_attachment=False,
                        download_name=filename
                    )
                else:
                    return jsonify({'success': False, 'error': 'Файл заявки не найден'}), 404
            
            elif request.method == 'POST':
                if 'requestFile' not in request.files:
                    return jsonify({'success': False, 'error': 'Файл не передан'}), 400
                
                file = request.files['requestFile']
                if file.filename == '':
                    return jsonify({'success': False, 'error': 'Файл не выбран'}), 400
                
                success, result = supabase_service.save_request_file(draft_id, file, extract_text_from_file)
                
                if success:
                    return jsonify({
                        'success': True,
                        'message': 'Файл заявки сохранен',
                        'filename': result
                    })
                else:
                    return jsonify({'success': False, 'error': result}), 400
            
            elif request.method == 'DELETE':
                success = supabase_service.delete_request_file(draft_id)
                
                if success:
                    return jsonify({'success': True, 'message': 'Файл заявки удален'})
                else:
                    return jsonify({'success': False, 'error': 'Ошибка удаления файла'}), 500
        except Exception as e:
            logger.error(f"Ошибка управления файлом заявки: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # ========== PROTOCOL GENERATION ==========
    @app.route('/api/generate-protocol', methods=['POST'])
    def generate_protocol():
        """Генерирует полный пакет: протокол Excel + все изображения"""
        try:
            from openpyxl import load_workbook
            import zipfile
            import tempfile
            
            if not os.path.exists(Config.TEMPLATE_PATH):
                return jsonify({'success': False, 'error': f'Шаблон {Config.TEMPLATE_PATH} не найден'}), 404
            
            data = request.form
            draft_id = data.get('draft_id')
            
            with tempfile.TemporaryDirectory() as temp_dir:
                folder_name = generate_folder_name(data)
                protocol_filename = generate_protocol_filename(data)
                
                # Генерируем протокол Excel
                wb = load_workbook(Config.TEMPLATE_PATH)
                ws = wb.active
                
                mapping = {
                    'workType': 'I1', 'machineType': 'C3', 'liftingCapacity': 'D3',
                    'serialNumber': 'J3', 'driveType': 'C5', 'driveNumber': 'F5',
                    'brakeResistor': 'E7', 'resistorCount': 'H7', 'electricMotor': 'D9',
                    'motorNumber': 'G9', 'EnginePower': 'K9', 'angleSensor': 'D11',
                    'angleSensorNumber': 'G11', 'speedSensorNumber': 'K11',
                    'leftVibrationSensor': 'D15', 'leftSensitivity': 'G15',
                    'leftSensorNumber': 'I15', 'rightVibrationSensor': 'D16',
                    'rightSensitivity': 'G16', 'rightSensorNumber': 'I16',
                    'measuringDevice': 'E18', 'measuringDeviceNumber': 'G18',
                    'signalProcessor': 'E20', 'signalProcessorNumber': 'G20',
                    'notes': 'A37'
                }
                
                for field, cell in mapping.items():
                    if field in data and data[field]:
                        ws[cell] = data[field]
                
                protocol_path = os.path.join(temp_dir, protocol_filename)
                wb.save(protocol_path)
                wb.close()
                
                # Создаем папку для изображений
                images_dir = os.path.join(temp_dir, f"{folder_name}_images")
                os.makedirs(images_dir, exist_ok=True)
                
                # Сохраняем изображения из формы
                images = request.files.getlist('images')
                saved_images = []
                
                for img in images:
                    if img and allowed_file(img.filename, Config.ALLOWED_EXTENSIONS):
                        filename = __import__('werkzeug.utils').secure_filename(img.filename)
                        img.seek(0)
                        img_data = img.read()
                        
                        compressed_data = compress_image(img_data, Config.IMAGE_MAX_SIZE, Config.IMAGE_QUALITY)
                        
                        img_path = os.path.join(images_dir, filename)
                        with open(img_path, 'wb') as f:
                            f.write(compressed_data)
                        saved_images.append(filename)
                
                # Загружаем дополнительные изображения из БД
                if draft_id:
                    draft = supabase_service.get_draft(draft_id)
                    if draft and draft.get('image_files'):
                        for img_filename in draft['image_files']:
                            if img_filename not in saved_images:
                                image_data = supabase_service.get_image(draft_id, img_filename)
                                if image_data:
                                    img_bytes, _ = image_data
                                    img_path = os.path.join(images_dir, img_filename)
                                    with open(img_path, 'wb') as f:
                                        f.write(img_bytes)
                                    saved_images.append(img_filename)
                
                # Создаем информационный файл
                info_content = f"""
ПРОТОКОЛ
=====================================
Тип станка: {data.get('machineType', 'Не указан')}
Грузоподъемность: {data.get('liftingCapacity', 'Не указана')}
Заводской номер: {data.get('serialNumber', 'Не указан')}
Вид работ: {data.get('workType', 'Не указан')}
Заказчик: {data.get('customer', 'Не указан')}
Статус: {data.get('machineStatus', 'Сборка')}
Плановая отгрузка: {data.get('shippingDate', 'Не указана')}
=====================================
"""
                
                info_path = os.path.join(temp_dir, "README.txt")
                with open(info_path, 'w', encoding='utf-8') as f:
                    f.write(info_content.strip())
                
                # Создаем ZIP архив
                zip_filename = generate_zip_filename(data)
                zip_path = os.path.join(temp_dir, zip_filename)
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(protocol_path, protocol_filename)
                    zf.write(info_path, "README.txt")
                    for img_file in os.listdir(images_dir):
                        img_full_path = os.path.join(images_dir, img_file)
                        zf.write(img_full_path, f"images/{img_file}")
                
                with open(zip_path, 'rb') as f:
                    zip_data = f.read()
                
                return send_file(
                    __import__('io').BytesIO(zip_data),
                    mimetype='application/zip',
                    as_attachment=True,
                    download_name=zip_filename
                )
        except Exception as e:
            logger.error(f"Ошибка генерации пакета: {e}")
            return jsonify({'success': False, 'error': f'Ошибка генерации протокола: {str(e)}'}), 500
    
    # ========== UPDATES ==========
    @app.route('/api/updates', methods=['GET'])
    def get_updates():
        """Получает все прошивки и утилиты"""
        try:
            section = request.args.get('section')
            result, error = supabase_service.get_updates(section)
            
            if error:
                return jsonify({'success': False, 'error': error}), 500
            
            return jsonify({
                'success': True,
                'updates': result['updates'],
                'last_update': result['last_update']
            })
        except Exception as e:
            logger.error(f"Ошибка получения прошивок: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/updates', methods=['POST'])
    def create_update():
        """Создает новую запись о прошивке"""
        try:
            data = request.get_json()
            
            if not data:
                return jsonify({'success': False, 'error': 'Нет данных для сохранения'}), 400
            
            required_fields = ['section', 'title', 'file']
            for field in required_fields:
                if field not in data or not data[field]:
                    return jsonify({'success': False, 'error': f'Поле {field} обязательно'}), 400
            
            success, error, update = supabase_service.create_update(data)
            
            if success:
                return jsonify({'success': True, 'message': 'Запись успешно создана', 'update': update})
            else:
                return jsonify({'success': False, 'error': error}), 400
        except Exception as e:
            logger.error(f"Ошибка создания записи: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/updates/<int:update_id>', methods=['PUT'])
    def update_update(update_id):
        """Обновляет существующую запись"""
        try:
            data = request.get_json()
            
            if not data:
                return jsonify({'success': False, 'error': 'Нет данных для обновления'}), 400
            
            success, error, update = supabase_service.update_update(update_id, data)
            
            if success:
                return jsonify({'success': True, 'message': 'Запись успешно обновлена', 'update': update})
            else:
                return jsonify({'success': False, 'error': error}), 400
        except Exception as e:
            logger.error(f"Ошибка обновления записи: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/updates/<int:update_id>', methods=['DELETE'])
    def delete_update(update_id):
        """Удаляет запись"""
        try:
            success, error = supabase_service.delete_update(update_id)
            
            if success:
                return jsonify({'success': True, 'message': 'Запись успешно удалена'})
            else:
                return jsonify({'success': False, 'error': error}), 500
        except Exception as e:
            logger.error(f"Ошибка удаления записи: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/updates/reorder', methods=['POST'])
    def reorder_updates():
        """Изменяет порядок сортировки записей"""
        try:
            data = request.get_json()
            orders = data.get('orders', [])
            
            if not orders:
                return jsonify({'success': False, 'error': 'Нет данных для сортировки'}), 400
            
            success, error = supabase_service.reorder_updates(orders)
            
            if success:
                return jsonify({'success': True, 'message': 'Порядок сортировки обновлен'})
            else:
                return jsonify({'success': False, 'error': error}), 500
        except Exception as e:
            logger.error(f"Ошибка обновления сортировки: {e}")
            return jsonify({'success': False, 'error': str(e)}), 500
    
    # ========== STATIC FILES ==========
    @app.route('/')
    def serve_main():
        """Главная страница"""
        return send_from_directory(app.static_folder, 'main.html')
    
    @app.route('/add-draft')
    def serve_add_draft():
        """Страница добавления станка"""
        return send_from_directory(app.static_folder, 'add_draft.html')
    
    @app.route('/view-machine.html')
    def serve_view_machine():
        """Страница просмотра станка"""
        return send_from_directory(app.static_folder, 'view_machine.html')
    
    @app.route('/static/<path:path>')
    def serve_static(path):
        """Статические файлы"""
        return send_from_directory('static', path)
    
    @app.route('/favicon.ico')
    def favicon():
        """Иконка сайта"""
        return send_from_directory('static', 'favicon.ico'), 404


def register_error_handlers(app):
    """Регистрация обработчиков ошибок"""
    
    @app.errorhandler(404)
    def not_found(error):
        return jsonify({
            'success': False,
            'error': 'Эндпоинт не найден'
        }), 404
    
    @app.errorhandler(500)
    def internal_error(error):
        app.logger.error(f"Внутренняя ошибка: {error}")
        return jsonify({
            'success': False,
            'error': 'Внутренняя ошибка сервера'
        }), 500


# Точка входа для gunicorn и других WSGI серверов
app = create_app()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(host='0.0.0.0', port=port, debug=False)
